"""
Chat Endpoints
BI Chat with AI assistant
"""

from typing import Annotated, Dict, Any, Optional, List
from fastapi import APIRouter, Depends, HTTPException, Request, status
from fastapi.responses import ORJSONResponse, StreamingResponse, JSONResponse, FileResponse
from pydantic import BaseModel, Field
from pathlib import Path
from types import SimpleNamespace
import json
import asyncio
import logging
import sys
import re
import structlog
from urllib.parse import urlparse
import numpy as np
import pandas as pd
from decimal import Decimal
from datetime import datetime, date
from uuid import uuid4

# Import core dependencies
from backend.app.api.dependencies import get_current_active_user, get_token_from_header_or_query, issue_stream_token, get_db
from backend.app.infrastructure.database.models import User
from backend.app.config.settings import settings
from backend.app.core.chat_capabilities import (
    get_chat_capabilities_for_user,
    get_chat_capability_diagnostics_for_user,
    require_chat_capability,
)
from backend.app.core.utils.response_cache import ResponseCache
from backend.app.core.utils.query_history import QueryHistory
from backend.app.core.utils.field_mapper import FieldMapper
from backend.app.core.rag.query_retriever import QueryRetriever
from backend.app.core.learning.pattern_matcher import PatternMatcher
# CodeGenAgent e CaculinhaBIAgent removidos - Arquitetura V2 deprecated
# Sistema agora usa ChatServiceV3 (Metrics-First)
from backend.app.core.llm_factory import LLMFactory, SmartLLM
from backend.app.core.utils.error_handler import APIError
from backend.app.core.utils.session_manager import SessionManager
from backend.app.core.utils.semantic_cache import cache_get, cache_set, cache_stats
from backend.app.core.utils.response_validator import validate_response, validator_stats
from backend.app.core.playground_mode import is_user_in_canary
from backend.app.core.security.content_safety import sanitize_citations, sanitize_text_label
from backend.app.core.utils.report_templates import get_official_report_templates
from backend.app.api.v1.endpoints.memory import get_memory_agent
# NEW SERVICE V3 - Metrics-First Architecture
from backend.app.services.chat_service_v3 import ChatServiceV3
from backend.app.services.chat_automation_service import ChatAutomationService
from backend.services.metrics import MetricsService
from sqlalchemy.ext.asyncio import AsyncSession
try:
    from backend.app.core.tools.competitive_intelligence_tool import pesquisar_precos_concorrentes
    from backend.app.core.tools.competitive_intelligence_tool import pesquisar_mercado_web
except (ImportError, OSError):
    pesquisar_precos_concorrentes = None
    pesquisar_mercado_web = None

logger = logging.getLogger(__name__)
trace_logger = structlog.get_logger("agentbi.chat.sse")

def _is_degraded_or_error_response(payload: Any) -> bool:
    """Avoid caching degraded/error messages and ignore them on cache reads."""
    text = str(payload).lower()
    degraded_markers = [
        "tempo limite",
        "quota estourada",
        "resource_exhausted",
        "serviço de ia temporariamente indisponível",
        "payload too large",
        "request too large",
        "erro ao processar",
        "busca externa não foi concluída nesta rodada",
        "nenhuma evidência pública validada",
    ]
    return any(marker in text for marker in degraded_markers)


def _sanitize_business_output(text: str) -> str:
    """Remove technical/internal leakage from final user-visible answer."""
    if not text:
        return text

    cleaned = str(text)
    blocked_patterns = [
        r"(?is)<\s*script\b[^>]*>.*?<\s*/\s*script\s*>",
        r"(?is)<\s*iframe\b[^>]*>.*?<\s*/\s*iframe\s*>",
        r"(?i)\bon\w+\s*=\s*(?:\"[^\"]*\"|'[^']*'|[^\s>]+)",
        r"(?i)\b(?:javascript|vbscript)\s*:",
        r"(?i)\bbase\s*=\s*[a-z0-9_.-]+",
        r"(?i)\b(parquet|duckdb|schema|coluna[s]?\s+internas?|tabela[s]?\s+internas?)\b",
        # Bloqueia caminhos de filesystem, mas preserva rotas web (ex.: /api/v1/...).
        r"(?i)\b[a-z]:\\[^\s]+",
        r"(?i)/(?:home|users?|usr|var|opt|etc|tmp|mnt|srv|root|proc|sys|dev|workspace|projects?|repos?)/[^\s]+",
        r"(?i)\b(venda_30dd|estoque_une|liquido_38|nomesegmento)\b",
    ]
    for pattern in blocked_patterns:
        cleaned = re.sub(pattern, "", cleaned)

    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


def _sanitize_response_for_role(text: str, role: str) -> str:
    """
    Sanitização adicional por perfil para evitar exposição de dados internos
    em respostas destinadas a usuários não privilegiados.
    """
    cleaned = _sanitize_business_output(text)
    role_norm = str(role or "").strip().lower()

    # Remove blocos técnicos/auditoria em qualquer perfil de chat.
    cleaned = re.sub(
        r"(?ims)^\s*##\s*SQL/Python\s*\n.*?(?=^\s*##\s+|\Z)",
        "",
        cleaned,
    )
    cleaned = re.sub(
        r"(?ims)^\s*##\s*Recorte e evid[êe]ncia\s*\n.*?(?=^\s*##\s+|\Z)",
        "",
        cleaned,
    )
    cleaned = re.sub(r"(?im)^\s*-\s*Template oficial:.*$", "", cleaned)
    cleaned = re.sub(r"(?im)^\s*Fonte:\s*.*$", "", cleaned)
    cleaned = re.sub(r"(?im)^\s*Confianca:\s*.*$", "", cleaned)
    cleaned = re.sub(r"(?im)^\s*Confiança:\s*.*$", "", cleaned)
    cleaned = re.sub(r"(?im)^\s*Citacoes:\s*.*$", "", cleaned)
    cleaned = re.sub(r"(?im)^\s*Citações:\s*.*$", "", cleaned)
    cleaned = re.sub(r"(?im)^\s*##\s*(Ação recomendada|Acao recomendada)\s*$", "## Próximas ações", cleaned)

    # Perfis restritos: remover detalhamento por loja/UNE.
    if role_norm not in {"admin", "analyst"}:
        table_pattern = r"(?ims)^(\s*##\s*Tabela operacional\s*\n)(.*?)(?=^\s*##\s+|\Z)"
        table_match = re.search(table_pattern, cleaned)
        if table_match:
            table_body = table_match.group(2)
            if re.search(r"(?i)\bUNE\b|Loja\s*\(UNE\)", table_body):
                cleaned = re.sub(
                    table_pattern,
                    "## Tabela operacional\n- Detalhamento por loja/UNE restrito para este perfil.\n\n",
                    cleaned,
                )
        cleaned = re.sub(r"(?i)UNE\s+l[ií]der:\s*\d+", "UNE lider: [restrito]", cleaned)

    cleaned = re.sub(r"\n{3,}", "\n\n", cleaned)
    return cleaned.strip()


def _is_competitive_market_query(query: str) -> bool:
    q = (query or "").lower()
    markers = [
        "concorrente", "concorrência", "cotação", "cotacao",
        "pesquisa de preço", "pesquisa de preco",
        "pesquisa de mercado", "preço de mercado", "preco de mercado",
        "benchmark de mercado", "pesquisa concorrencial",
        "comparar preço", "comparar preco",
        "americanas", "kalunga", "bellart", "shopee", "amazon", "mercado livre",
        "google shopping",
    ]
    return any(m in q for m in markers)


def _is_chat_allowed_for_user(user: User) -> bool:
    return is_user_in_canary(
        user_id=str(getattr(user, "id", "") or ""),
        username=str(getattr(user, "username", "") or ""),
        role=str(getattr(user, "role", "") or ""),
        canary_enabled=settings.CHAT_CANARY_ENABLED,
        allowed_roles_csv=settings.CHAT_CANARY_ALLOWED_ROLES,
        allowed_users_csv=settings.CHAT_CANARY_ALLOWED_USERS,
    )


def _has_specific_competitor(query: str) -> bool:
    """Retorna True se a query menciona um concorrente específico pelo nome."""
    q = (query or "").lower()
    competitors = [
        "americanas", "kalunga", "bellart", "shopee", "amazon",
        "casa&video", "casa e video", "le biscuit", "lebiscuit",
        "tubarão", "tubarao", "tid", "amigão", "amigao",
        "mercado livre", "mercadolivre", "meli",
    ]
    return any(c in q for c in competitors)


def _should_use_market_web_fast_path(query: str) -> bool:
    """
    Regras de roteamento para pesquisa de mercado:
    - market_web: pesquisa aberta genérica ou pedido explícito de Mercado Livre.
      Hoje essa ferramenta já consulta múltiplos providers públicos.
    - competitive_research: concorrente específico (fora ML) ou pedido
      explícito para "concorrentes".
    """
    q = (query or "").lower()
    ml_aliases = ["mercado livre", "mercadolivre", "meli"]
    other_competitors = [
        "americanas", "kalunga", "bellart", "shopee", "amazon",
        "casa&video", "casa e video", "le biscuit", "lebiscuit",
        "tubarão", "tubarao", "tid", "amigão", "amigao",
    ]
    competitor_terms = [
        "concorrente",
        "concorrentes",
        "concorrência",
        "concorrencia",
    ]
    mentions_ml = any(k in q for k in ml_aliases)
    mentions_other_competitor = any(k in q for k in other_competitors)
    mentions_competitor_terms = any(k in q for k in competitor_terms)

    # Mercado Livre explícito sem outro concorrente -> mercado aberto.
    if mentions_ml and not mentions_other_competitor:
        return True

    # Concorrente específico (fora ML) -> pesquisa concorrencial.
    if mentions_other_competitor:
        return False

    # Pedido explícito de concorrentes sem nomes -> pesquisa concorrencial.
    if mentions_competitor_terms:
        return False

    # Pesquisa de mercado genérica -> mercado web aberto multi-provider.
    return True


def _extract_market_product_hint(query: str) -> str:
    q = (query or "").strip()
    if not q:
        return "item solicitado"

    lowered = q.lower()
    lowered = re.sub(r"^(faca|faça|faz|fazer)\s+(uma\s+)?", "", lowered)
    lowered = re.sub(r"^(realize|realizar|realiza)\s+(uma\s+)?", "", lowered)
    lowered = re.sub(r"^(pesquisa|pesquise|compare|comparar|benchmark)\s+", "", lowered)
    lowered = re.sub(r"^(de\s+mercado|de\s+pre[çc]o)\s+", "", lowered)
    lowered = re.sub(r"^(do|da|de|o|a)?\s*produto\s+", "", lowered)
    lowered = re.sub(
        r"\b(nos?\s+concorrentes?\s+.+)$",
        "",
        lowered,
    )
    lowered = re.sub(
        r"\b(?:na|no|em)\s+(mercado livre|mercadolivre|meli|kalunga|americanas|amazon|shopee|le biscuit|lebiscuit|casa&video|casa e video|bellart|amig[aã]o|tubar[aã]o|tid'?s?)\b.*$",
        "",
        lowered,
    )
    lowered = re.sub(
        r"\b(?:em|no estado)\s+(rj|rio de janeiro|mg|minas gerais|es|esp[ií]rito santo|espirito santo)\b.*$",
        "",
        lowered,
    )
    lowered = re.sub(r"\s+", " ", lowered).strip(" .,-")
    return lowered or "item solicitado"


def _competitive_timeout_business_message(query: str) -> str:
    produto = _extract_market_product_hint(query)
    return (
        "## Resumo executivo\n"
        f"- A busca de mercado para {produto} foi concluída parcialmente nesta rodada.\n"
        "- Não houve evidência pública suficiente para consolidar preço confiável agora.\n\n"
        "## Tabela operacional\n"
        "| Indicador | Valor |\n"
        "|---|---|\n"
        "| Situação da busca | Parcial |\n"
        "| Evidência pública confiável | Insuficiente |\n\n"
        "## Próximas ações\n"
        "- Use cotação direta com 2-3 fornecedores para decisão imediata.\n"
        "- Refaça a pesquisa com SKU/marca e especificação completa para ampliar cobertura.\n\n"
        "## Próxima consulta sugerida\n"
        f"- \"pesquisa de mercado de {produto} marca X, medida Y, em RJ\""
    )


def _extract_market_state(query: str) -> str:
    q = (query or "").lower()
    if re.search(r"\b(rj|rio de janeiro)\b", q):
        return "RJ"
    if re.search(r"\b(mg|minas gerais)\b", q):
        return "MG"
    if re.search(r"\b(es|esp[ií]rito santo|espirito santo)\b", q):
        return "ES"
    return "RJ"


def _extract_market_competitors_csv(query: str) -> str:
    q = (query or "").lower()
    mapping = [
        ("americanas", ["americanas", "lojas americanas"]),
        ("kalunga", ["kalunga"]),
        ("bellart", ["bellart"]),
        ("amazon", ["amazon"]),
        ("shopee", ["shopee"]),
        ("mercado livre", ["mercado livre", "mercadolivre", "meli"]),
        ("casa&video", ["casa&video", "casa e video", "casaevideo"]),
        ("le biscuit", ["le biscuit", "lebiscuit"]),
    ]
    found = []
    for canonical, aliases in mapping:
        if any(alias in q for alias in aliases):
            found.append(canonical)
    return ",".join(found)


def _extract_market_followup_context(chat_history: List[Dict[str, Any]]) -> Dict[str, Any]:
    context: Dict[str, Any] = {}
    if not isinstance(chat_history, list):
        return context

    last_user_query = None
    last_assistant_msg = None
    for msg in reversed(chat_history):
        role = str(msg.get("role", "")).lower()
        if role == "assistant" and last_assistant_msg is None:
            last_assistant_msg = msg
        elif role == "user" and last_user_query is None:
            content = str(msg.get("content", "")).strip()
            if content:
                last_user_query = content
        if last_user_query and last_assistant_msg is not None:
            break

    if last_user_query:
        context["last_user_query"] = last_user_query
        context["market_product_hint"] = _extract_market_product_hint(last_user_query)

    if isinstance(last_assistant_msg, dict):
        assistant_content = str(last_assistant_msg.get("content", "")).strip()
        if assistant_content:
            context["last_assistant_content"] = assistant_content

        metadata = last_assistant_msg.get("metadata")
        if isinstance(metadata, dict):
            meta_context = metadata.get("context")
            if isinstance(meta_context, dict):
                for key in ("source", "market_product_hint", "market_competitors"):
                    if meta_context.get(key) not in (None, "", []):
                        context[key] = meta_context.get(key)
            if metadata.get("source") not in (None, "", []):
                context["source"] = metadata.get("source")

    return context


def _resolve_competitive_market_followup_query(query: str, chat_history: List[Dict[str, Any]]) -> str:
    q = str(query or "").strip()
    if not q:
        return q

    lowered = q.lower()
    if not _is_competitive_market_query(lowered):
        return q

    followup_context = _extract_market_followup_context(chat_history)
    source = str(followup_context.get("source") or "").lower()
    last_assistant = str(followup_context.get("last_assistant_content") or "").lower()
    has_market_context = (
        "pesquisar_precos_concorrentes" in source
        or "pesquisar_mercado_web" in source
        or "context.market_research_followup" in source
        or "pesquisa de mercado" in last_assistant
        or "pesquisa concorrencial" in last_assistant
    )
    if not has_market_context:
        return q

    word_count = len([token for token in lowered.split() if token.strip()])
    is_short_followup = word_count <= 8 or lowered.startswith(("e ", "na ", "no "))
    has_context_marker = any(
        marker in lowered
        for marker in (
            "com base",
            "nessa pesquisa",
            "nesse benchmark",
            "anterior",
            "ultima resposta",
            "última resposta",
        )
    )
    if not (is_short_followup or has_context_marker):
        return q

    market_product = str(followup_context.get("market_product_hint") or "").strip()
    if not market_product:
        return q

    competitors_csv = _extract_market_competitors_csv(lowered)
    if any(alias in lowered for alias in ("mercado livre", "mercadolivre", "meli")) and not competitors_csv.replace("mercado livre", "").strip(", "):
        return f"pesquisa de mercado de {market_product} no mercado livre"
    if competitors_csv:
        return f"pesquisa de mercado de {market_product} nos concorrentes {competitors_csv.replace(',', ', ')}"
    return f"pesquisa de mercado de {market_product}"


def _price_to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    txt = str(value).strip()
    if not txt:
        return None
    txt = txt.replace("R$", "").replace(" ", "")
    if "," in txt:
        txt = txt.replace(".", "").replace(",", ".")
    try:
        return float(txt)
    except Exception:
        return None


def _format_brl(value: Optional[float]) -> str:
    if value is None:
        return "-"
    return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def _business_source_label(item: Dict[str, Any]) -> str:
    fonte = str(item.get("fonte") or "").strip().lower()
    if fonte in {"serpapi_shopping", "serpapi_google_shopping"}:
        competitor = str(item.get("concorrente") or "").strip()
        if competitor:
            return competitor

    url = str(item.get("url") or "").strip()
    if url:
        domain = urlparse(url).netloc.lower().replace("www.", "")
        if domain:
            return domain
    if fonte in {"manual", "benchmark_seed_local", "base_manual_concorrencial", "csv_compras"}:
        return ""
    return fonte or "fonte_publica"


def _is_public_price_evidence(item: Dict[str, Any]) -> bool:
    if not isinstance(item, dict):
        return False
    price = _price_to_float(item.get("preco"))
    if price is None:
        return False
    url = str(item.get("url") or "").strip()
    if not url:
        return False
    source = str(item.get("fonte") or "").strip().lower()
    if source in {"manual", "benchmark_seed_local", "base_manual_concorrencial", "csv_compras"}:
        return False
    domain = urlparse(url).netloc.lower().replace("www.", "")
    if not domain or domain == "manual":
        return False
    return True


# ---------------------------------------------------------------------------
# Preço interno do parquet + cache de resultados para download
# ---------------------------------------------------------------------------
import io
import tempfile
from uuid import uuid4 as _uuid4

# Cache simples em memória: {search_id: {"rows": [...], "produto": str, "internal_price": {...}, "created_at": str}}
_market_search_cache: Dict[str, Dict[str, Any]] = {}


def _lookup_internal_price(product_hint: str) -> Dict[str, Any]:
    """Busca preço de venda (LIQUIDO_38) e custo (ULTIMA_ENTRADA_CUSTO_CD) no parquet."""
    try:
        import duckdb
        parquet_path = str(Path(settings.PARQUET_DATA_PATH).resolve())
        if not Path(parquet_path).exists():
            # Tentar caminho relativo ao backend
            parquet_path = str(Path(__file__).resolve().parent.parent.parent.parent / settings.PARQUET_DATA_PATH)
        if not Path(parquet_path).exists():
            return {}

        terms = [t.strip().lower() for t in str(product_hint or "").split() if t.strip()]
        terms = terms[:4]  # Máximo 4 termos

        if not terms:
            return {}

        sql = (
            "SELECT "
            "  NOME, "
            "  TRY_CAST(REPLACE(COALESCE(LIQUIDO_38, ''), ',', '.') AS DOUBLE) AS LIQUIDO_38_NUM, "
            "  TRY_CAST(REPLACE(COALESCE(ULTIMA_ENTRADA_CUSTO_CD, ''), ',', '.') AS DOUBLE) AS ULTIMA_ENTRADA_CUSTO_CD_NUM "
            "FROM read_parquet(?) "
            "WHERE TRY_CAST(REPLACE(COALESCE(LIQUIDO_38, ''), ',', '.') AS DOUBLE) > 0 "
        )
        params: List[Any] = [parquet_path]
        for term in terms:
            sql += "AND LOWER(COALESCE(NOME, '')) LIKE ? "
            params.append(f"%{term}%")
        sql += "LIMIT 5"

        conn = duckdb.connect()
        try:
            result = conn.execute(sql, params).fetchall()
        finally:
            conn.close()

        if not result:
            return {}

        # Pegar a média dos preços encontrados
        precos_venda = [float(r[1]) for r in result if r[1] and float(r[1]) > 0]
        custos = [float(r[2]) for r in result if r[2] and float(r[2]) > 0]
        nomes = [str(r[0]) for r in result]

        return {
            "preco_venda": round(sum(precos_venda) / len(precos_venda), 2) if precos_venda else None,
            "custo": round(sum(custos) / len(custos), 2) if custos else None,
            "produtos_encontrados": len(result),
            "nome_exemplo": nomes[0] if nomes else None,
        }
    except Exception as e:
        logger.warning(f"Erro ao buscar preço interno: {e}")
        return {}


def _competitive_no_evidence_business_message(query: str) -> str:
    produto = _extract_market_product_hint(query)
    return (
        "## Resumo executivo\n"
        f"- Nao encontrei nenhuma evidência pública validada para {produto} nesta consulta.\n\n"
        "## Como obter resultado mais preciso\n"
        "- Informe marca, gramatura/medida e embalagem (ex.: \"TNT branco 40g rolo 1,40m x 50m\").\n"
        "- Se quiser, inclua concorrente e estado (ex.: Kalunga em RJ).\n\n"
        "## Proxima consulta sugerida\n"
        f"- \"pesquisa de mercado de {produto} marca X medida Y em RJ\""
    )


def _build_competitive_structured_business_message(query: str, payload: Dict[str, Any]) -> str:
    if not isinstance(payload, dict) or payload.get("status") != "success":
        return _competitive_no_evidence_business_message(query)

    raw_items = payload.get("itens") or []
    if not isinstance(raw_items, list):
        raw_items = []

    valid_items = [item for item in raw_items if _is_public_price_evidence(item)]
    if not valid_items:
        return _competitive_no_evidence_business_message(query)

    rows = []
    for item in valid_items:
        price = _price_to_float(item.get("preco"))
        if price is None:
            continue
        rows.append(
            {
                "concorrente": str(item.get("concorrente") or "concorrente").strip(),
                "produto": str(item.get("produto") or "produto").strip(),
                "preco": price,
                "fonte": _business_source_label(item),
            }
        )

    if not rows:
        return _competitive_no_evidence_business_message(query)

    rows = sorted(rows, key=lambda r: r["preco"])
    avg_price = sum(r["preco"] for r in rows) / len(rows)
    min_row = rows[0]
    max_row = rows[-1]

    # Buscar preço interno
    produto = _extract_market_product_hint(query)
    internal = _lookup_internal_price(produto)

    # Salvar no cache para download posterior
    search_id = str(_uuid4())[:8]
    _market_search_cache[search_id] = {
        "rows": rows,
        "produto": produto,
        "internal_price": internal,
        "avg_price": avg_price,
        "created_at": datetime.now().isoformat(),
    }
    # Limpar cache antigo (manter apenas últimos 20)
    if len(_market_search_cache) > 20:
        oldest_key = next(iter(_market_search_cache))
        _market_search_cache.pop(oldest_key, None)

    table_lines = [
        "| Concorrente | Produto | Preco (R$) | Fonte |",
        "|---|---|---:|---|",
    ]
    for row in rows[:15]:
        table_lines.append(
            f"| {row['concorrente']} | {row['produto'][:60]} | {_format_brl(row['preco'])} | {row['fonte'] or 'site'} |"
        )
    competitors_found: List[str] = []
    for row in rows:
        comp = str(row.get("concorrente") or "").strip()
        if comp and comp not in competitors_found:
            competitors_found.append(comp)
    competitors_line = ", ".join(competitors_found[:6]) if competitors_found else "N/D"
    if len(competitors_found) > 6:
        competitors_line += ", ..."

    # Montar seção de preço interno
    internal_section = ""
    acao_interna = ""
    if internal.get("preco_venda"):
        pv = internal["preco_venda"]
        custo = internal.get("custo")
        diff_pct = ((pv - avg_price) / avg_price) * 100 if avg_price > 0 else 0
        position = "ACIMA" if diff_pct > 0 else "ABAIXO"

        internal_section = f"- Nosso preco de venda: R$ {_format_brl(pv)}"
        if custo:
            internal_section += f" | Nosso custo: R$ {_format_brl(custo)}"
        internal_section += "\n"
        
        acao_interna = f"- Nosso preco esta {abs(diff_pct):.0f}% {position} da media de mercado (R$ {_format_brl(pv)} vs R$ {_format_brl(avg_price)}).\n"
        if custo:
            margem = ((pv - custo) / pv) * 100 if pv > 0 else 0
            acao_interna += f"- Margem estimada: {margem:.0f}% sobre custo.\n"

    # Links relativos para funcionar em qualquer ambiente (dev/prod) via mesmo host do frontend.
    download_section = (
        f"\n## Download dos resultados\n"
        f"- [Baixar Excel (.xlsx)](/api/v1/chat/market-research/download/{search_id}?format=xlsx)\n"
        f"- [Baixar CSV (.csv)](/api/v1/chat/market-research/download/{search_id}?format=csv)\n"
    )

    action_block = (
        f"- Use a faixa entre R$ {_format_brl(min_row['preco'])} e R$ {_format_brl(avg_price)} como referencia de negociacao.\n"
        if not internal.get("preco_venda")
        else acao_interna
    )
    if len(competitors_found) == 1:
        action_block += (
            f"- Cobertura concentrada em {competitors_found[0]}; "
            "para ampliar a comparação, informe concorrentes-alvo (ex.: Kalunga, Americanas, Shopee).\n"
        )

    return (
        "## Resumo executivo\n"
        + f"- Pesquisa de mercado concluida para **{produto}** com {len(rows)} referencias.\n"
        + f"- Concorrentes com preço identificado: {competitors_line}\n"
        + (f"{internal_section}" if internal_section else "")
        + f"- Preco medio de mercado: R$ {_format_brl(avg_price)}\n"
        + f"- Faixa: R$ {_format_brl(min_row['preco'])} ate R$ {_format_brl(max_row['preco'])}\n"
        + "## Tabela operacional\n"
        + "\n".join(table_lines)
        + "\n\n## Próximas ações\n"
        + action_block
        + download_section
    )


def _payload_has_public_price_evidence(payload: Dict[str, Any]) -> bool:
    if not isinstance(payload, dict):
        return False
    items = payload.get("itens") or []
    if not isinstance(items, list):
        return False
    return any(_is_public_price_evidence(item) for item in items if isinstance(item, dict))


def _market_fast_path_timeout_seconds() -> float:
    raw = getattr(settings, "MARKET_FAST_PATH_TIMEOUT_SEC", 40) or 40
    try:
        timeout = float(raw)
    except (TypeError, ValueError):
        timeout = 40.0
    return max(10.0, min(timeout, 90.0))


def _market_contract_from_payload(payload: Dict[str, Any], default_source: str) -> Dict[str, Any]:
    data = payload if isinstance(payload, dict) else {}
    source = data.get("source") or default_source
    confidence = data.get("confidence")
    mode = data.get("mode") or "deterministic_tool"
    citations = data.get("citations")

    try:
        confidence_value = float(confidence) if confidence is not None else None
    except (TypeError, ValueError):
        confidence_value = None
    if confidence_value is None:
        total_items = int(data.get("total_itens", 0) or 0)
        confidence_value = round(max(0.1, min(0.9, 0.3 + total_items * 0.04)), 2)

    if not isinstance(citations, list):
        citations = []
    if not citations:
        fontes = data.get("fontes_consultadas", [])
        if isinstance(fontes, list):
            derived: List[Dict[str, Any]] = []
            for src in fontes[:8]:
                if not isinstance(src, dict):
                    continue
                derived.append(
                    {
                        "source": str(src.get("fonte") or "fonte_publica"),
                        "domain": str(src.get("dominio") or "n/a"),
                        "url": str(src.get("url") or "").strip(),
                        "competitor": str(src.get("concorrente") or "n/a"),
                    }
                )
            citations = derived

    citations = sanitize_citations(citations)
    return {
        "source": source,
        "confidence": confidence_value,
        "mode": mode,
        "citations": citations,
    }


async def _run_competitive_market_fast_path(query: str, return_payload: bool = False) -> Any:
    async def _try_market_web_fallback(reason: str) -> Any:
        if pesquisar_mercado_web is None or _has_specific_competitor(query):
            return None
        logger.info("Competitive fast-path %s; aplicando fallback market_web.", reason)
        try:
            try:
                fallback_result = await _run_market_research_fast_path(query, return_payload=return_payload)
            except TypeError:
                # Compatibilidade com testes/mocks legados sem o novo parâmetro.
                fallback_result = await _run_market_research_fast_path(query)
        except Exception as exc:
            logger.warning("Fallback market_web falhou apos %s: %s", reason, exc)
            return None

        if isinstance(fallback_result, dict):
            fallback_text = str(fallback_result.get("text") or "").strip()
            if fallback_text:
                return fallback_result
            return None
        if isinstance(fallback_result, str) and fallback_result.strip():
            return fallback_result
        return None

    if pesquisar_precos_concorrentes is None:
        fallback_result = await _try_market_web_fallback("indisponibilidade_competitive_tool")
        if fallback_result is not None:
            return fallback_result
        fallback_text = _competitive_no_evidence_business_message(query)
        if return_payload:
            return {
                "text": fallback_text,
                "payload": {
                    "source": "tool.pesquisar_precos_concorrentes",
                    "confidence": 0.1,
                    "mode": "deterministic_no_evidence",
                    "citations": [],
                },
            }
        return fallback_text

    estado = _extract_market_state(query)
    concorrentes_csv = _extract_market_competitors_csv(query)

    def _invoke_tool() -> Any:
        tool_input = {
            "descricao_produto": query,
            "segmento": "",
            "estado": estado,
            "cidade": "",
            "limite": "12",
            "concorrentes": concorrentes_csv,
        }
        if hasattr(pesquisar_precos_concorrentes, "invoke"):
            return pesquisar_precos_concorrentes.invoke(tool_input)
        return pesquisar_precos_concorrentes(**tool_input)

    try:
        raw = await asyncio.wait_for(asyncio.to_thread(_invoke_tool), timeout=_market_fast_path_timeout_seconds())
    except asyncio.TimeoutError:
        logger.warning("Competitive fast-path timeout para query: %s", query)
        fallback_result = await _try_market_web_fallback("timeout")
        if fallback_result is not None:
            return fallback_result
        fallback_text = _competitive_no_evidence_business_message(query)
        if return_payload:
            return {
                "text": fallback_text,
                "payload": {
                    "source": "tool.pesquisar_precos_concorrentes",
                    "confidence": 0.12,
                    "mode": "deterministic_degraded_timeout",
                    "citations": [],
                },
            }
        return fallback_text
    except Exception as exc:
        logger.error("Competitive fast-path error: %s", exc, exc_info=True)
        fallback_result = await _try_market_web_fallback("erro")
        if fallback_result is not None:
            return fallback_result
        fallback_text = _competitive_no_evidence_business_message(query)
        if return_payload:
            return {
                "text": fallback_text,
                "payload": {
                    "source": "tool.pesquisar_precos_concorrentes",
                    "confidence": 0.1,
                    "mode": "deterministic_error",
                    "citations": [],
                },
            }
        return fallback_text

    payload: Dict[str, Any]
    if isinstance(raw, dict):
        payload = raw
    elif isinstance(raw, str):
        try:
            parsed = json.loads(raw)
            payload = parsed if isinstance(parsed, dict) else {"status": "error"}
        except Exception:
            payload = {"status": "error"}
    else:
        payload = {"status": "error"}

    # Fallback de robustez: quando pesquisa concorrencial não traz evidência pública
    # validada, tenta pesquisa de mercado aberta para não perder cobertura.
    if (
        not _payload_has_public_price_evidence(payload)
        and pesquisar_mercado_web is not None
        and not _has_specific_competitor(query)
    ):
        fallback_result = await _try_market_web_fallback("sem_evidencia_publica")
        if fallback_result is not None:
            return fallback_result

    text = _build_competitive_structured_business_message(query, payload)
    if return_payload:
        return {"text": text, "payload": payload}
    return text


async def _run_market_research_fast_path(query: str, return_payload: bool = False) -> Any:
    """Fast path para pesquisa de mercado genérica usando pesquisar_mercado_web."""
    if pesquisar_mercado_web is None:
        fallback_text = _competitive_no_evidence_business_message(query)
        if return_payload:
            return {
                "text": fallback_text,
                "payload": {
                    "source": "tool.pesquisar_mercado_web",
                    "confidence": 0.1,
                    "mode": "deterministic_no_evidence",
                    "citations": [],
                },
            }
        return fallback_text

    produto = _extract_market_product_hint(query)

    def _invoke_tool() -> Any:
        fn = getattr(pesquisar_mercado_web, "func", pesquisar_mercado_web)
        return fn(termo_pesquisa=produto, limite="15")

    try:
        raw = await asyncio.wait_for(asyncio.to_thread(_invoke_tool), timeout=_market_fast_path_timeout_seconds())
    except asyncio.TimeoutError:
        logger.warning("Market research fast-path timeout para query: %s", query)
        fallback_text = _competitive_no_evidence_business_message(query)
        if return_payload:
            return {
                "text": fallback_text,
                "payload": {
                    "source": "tool.pesquisar_mercado_web",
                    "confidence": 0.12,
                    "mode": "deterministic_degraded_timeout",
                    "citations": [],
                },
            }
        return fallback_text
    except Exception as exc:
        logger.error("Market research fast-path error: %s", exc, exc_info=True)
        fallback_text = _competitive_no_evidence_business_message(query)
        if return_payload:
            return {
                "text": fallback_text,
                "payload": {
                    "source": "tool.pesquisar_mercado_web",
                    "confidence": 0.1,
                    "mode": "deterministic_error",
                    "citations": [],
                },
            }
        return fallback_text

    payload: Dict[str, Any]
    if isinstance(raw, dict):
        payload = raw
    elif isinstance(raw, str):
        try:
            parsed = json.loads(raw)
            payload = parsed if isinstance(parsed, dict) else {"status": "error"}
        except Exception:
            payload = {"status": "error"}
    else:
        payload = {"status": "error"}

    text = _build_competitive_structured_business_message(query, payload)
    if return_payload:
        return {"text": text, "payload": payload}
    return text


def safe_json_dumps(obj: Any, **kwargs) -> str:
    """
    Safely serialize any Python object to JSON string.
    Handles MapComposite, numpy types, pandas types, datetime, and other non-serializable objects.
    """
    def default_handler(o):
        # Handle numpy types
        if isinstance(o, (np.integer, np.int64, np.int32, np.int16, np.int8)):
            return int(o)
        elif isinstance(o, (np.floating, np.float64, np.float32, np.float16)):
            if np.isnan(o) or np.isinf(o):
                return None
            return float(o)
        elif isinstance(o, np.ndarray):
            return o.tolist()
        elif isinstance(o, np.bool_):
            return bool(o)

        # Handle pandas types
        elif isinstance(o, pd.Timestamp):
            return o.isoformat()
        elif isinstance(o, pd.Timedelta):
            return str(o)
        elif pd.isna(o):
            return None

        # Handle datetime types
        elif isinstance(o, (datetime, date)):
            return o.isoformat()

        # Handle Decimal
        elif isinstance(o, Decimal):
            return float(o)

        # Handle bytes
        elif isinstance(o, bytes):
            return o.decode('utf-8', errors='ignore')

        # Handle SQLAlchemy Row/MapComposite and similar mapping types
        elif hasattr(o, '_mapping'):
            return dict(o._mapping)
        elif hasattr(o, '__dict__') and not isinstance(o, type):
            # Generic object with __dict__
            return {k: v for k, v in o.__dict__.items() if not k.startswith('_')}

        # Last resort: convert to string
        else:
            return str(o)

    try:
        # Merge default handler with any custom kwargs
        if 'default' not in kwargs:
            kwargs['default'] = default_handler
        return json.dumps(obj, **kwargs)
    except Exception as e:
        logger.error(f"Failed to serialize object: {e}", exc_info=True)
        # Ultimate fallback: return error as JSON
        return json.dumps({"error": f"Serialization failed: {str(e)}"}, ensure_ascii=False)


# [OK] PERFORMANCE FIX: Initialization moved to startup background task
# This prevents the 15s delay on the first user query.
chat_service_v3 = None  # Metrics-First Architecture
session_manager = None
query_history = None  # [OK] FIX: Added missing variable for feedback endpoint
chat_automation_service = ChatAutomationService()
_init_lock = asyncio.Lock()
_CHAT_CACHE_VERSION = "chatbi_v3_20260308_r7"
_SSE_EVENT_POLL_TIMEOUT_SECONDS = 0.1
_SSE_KEEPALIVE_INTERVAL_TICKS = 50


def _build_stream_validation_context(query: str, payload: Dict[str, Any]) -> Dict[str, Any]:
    global chat_service_v3
    if chat_service_v3 is not None and hasattr(chat_service_v3, "_build_response_validation_context"):
        try:
            return chat_service_v3._build_response_validation_context(query, payload)
        except Exception as exc:
            logger.warning("stream_validation_context_failed: %s", exc)
    return {}


def _build_stream_validation_block_response(
    *,
    query: str,
    validation_result: Any,
    validation_context: Dict[str, Any],
) -> Dict[str, Any]:
    global chat_service_v3
    if chat_service_v3 is not None and hasattr(chat_service_v3, "_build_validation_block_response"):
        try:
            return chat_service_v3._build_validation_block_response(
                query=query,
                validation_result=validation_result,
                validation_context=validation_context,
            )
        except Exception as exc:
            logger.warning("stream_validation_block_build_failed: %s", exc)
    return {
        "type": "text",
        "result": {
            "mensagem": (
                "## Resumo executivo\n"
                "- A resposta gerada nesta rodada não passou na validação interna de coerência.\n\n"
                "## Tabela operacional\n"
                "- O sistema descartou a saída por incompatibilidade entre intenção e payload.\n\n"
                "## Próximas ações\n"
                "- Refaça a pergunta com o objetivo principal de forma direta."
            )
        },
        "source": "policy.response_validation",
        "mode": "validation_block",
        "confidence": 0.0,
    }


def _resolve_active_data_fingerprint() -> str:
    """Inclui assinatura do parquet ativo para invalidar cache após sync de base."""
    parquet_path_raw = str(
        getattr(settings, "PARQUET_DATA_PATH", None)
        or getattr(settings, "PARQUET_FILE_PATH", None)
        or ""
    ).strip()
    if not parquet_path_raw:
        return "dataset:unknown"

    parquet_path = Path(parquet_path_raw)
    try:
        parquet_path = parquet_path.resolve()
    except OSError:
        pass

    try:
        stat = parquet_path.stat()
        return f"parquet:{int(stat.st_mtime_ns)}:{int(stat.st_size)}"
    except OSError:
        return f"parquet:missing:{parquet_path.name}"


def _build_chat_cache_key(session_id: Optional[str], query: str) -> str:
    session_component = str(session_id or "anonymous")
    data_fingerprint = _resolve_active_data_fingerprint()
    return f"{_CHAT_CACHE_VERSION}:{data_fingerprint}:{session_component}:{query}"


_CHAT_MODE_LABELS = {
    "executive_overview": "resumo executivo",
    "sales_by_store": "vendas por loja",
    "critical_stock": "ruptura e reposicao",
    "promotion_margin": "promocao e margem",
    "market_benchmark": "benchmark de mercado",
    "seasonal_plan": "planejamento sazonal",
}

_CHAT_MODE_INSTRUCTIONS = {
    "executive_overview": [
        "responda como sumario executivo para decisao gerencial",
        "priorize 3 a 5 achados materialmente relevantes",
        "feche com acoes praticas e riscos imediatos",
    ],
    "sales_by_store": [
        "compare desempenho entre lojas ou UNEs",
        "destaque top performers, piores desempenhos e outliers",
        "se houver dados suficientes, sugira investigacao por mix, ruptura ou precificacao",
    ],
    "critical_stock": [
        "priorize ruptura, cobertura de estoque e urgencia de reposicao",
        "quantifique impacto potencial em venda perdida quando possivel",
        "sugira transferencia, compra ou ajuste de abastecimento com prioridade clara",
    ],
    "promotion_margin": [
        "trate margem como restricao obrigatoria",
        "avalie desconto, elasticidade esperada e risco de erosao de margem",
        "recomende promover, ajustar desconto ou abortar a acao com justificativa objetiva",
    ],
    "market_benchmark": [
        "compare preco interno, faixa de mercado e posicionamento competitivo",
        "destaque desvios relevantes versus concorrencia",
        "recomende ajuste de preco, manutencao ou monitoramento com base nos dados",
    ],
    "seasonal_plan": [
        "considere sazonalidade comercial e preparo antecipado",
        "projete demanda, estoque e risco de ruptura ou sobra",
        "responda com plano operacional por periodo e prioridade",
    ],
}


def _sanitize_playbook_context(raw_value: Any) -> Dict[str, str]:
    if not isinstance(raw_value, dict):
        return {}
    sanitized: Dict[str, str] = {}
    for key in ("product", "segment", "une", "period", "objective"):
        value = raw_value.get(key)
        text = str(value or "").strip()
        if text:
            sanitized[key] = text[:240]
    return sanitized


def _parse_playbook_context_param(raw_value: Optional[str]) -> Dict[str, str]:
    if raw_value in (None, ""):
        return {}
    try:
        decoded = json.loads(str(raw_value))
    except (TypeError, json.JSONDecodeError):
        return {}
    return _sanitize_playbook_context(decoded)


def _sanitize_guided_action(raw_value: Any) -> Dict[str, Any]:
    if not isinstance(raw_value, dict):
        return {}

    sanitized: Dict[str, Any] = {}
    for key in (
        "actionId",
        "actionLabel",
        "source",
        "playbookId",
        "prompt",
        "executionPolicy",
        "outputPreference",
        "missingDataBehavior",
    ):
        value = raw_value.get(key)
        text = str(value or "").strip()
        if text:
            sanitized[key] = text[:240]

    if isinstance(raw_value.get("directSend"), bool):
        sanitized["directSend"] = bool(raw_value["directSend"])

    tool_hints = raw_value.get("toolHints")
    if isinstance(tool_hints, list):
        normalized_hints = [
            str(item).strip()[:80]
            for item in tool_hints
            if str(item or "").strip()
        ]
        if normalized_hints:
            sanitized["toolHints"] = normalized_hints[:8]

    return sanitized


def _parse_guided_action_param(raw_value: Optional[str]) -> Dict[str, Any]:
    if raw_value in (None, ""):
        return {}
    try:
        decoded = json.loads(str(raw_value))
    except (TypeError, json.JSONDecodeError):
        return {}
    return _sanitize_guided_action(decoded)


def _build_guided_chat_query(
    query: str,
    chat_mode: Optional[str],
    playbook_context: Optional[Dict[str, Any]],
    guided_action: Optional[Dict[str, Any]] = None,
) -> str:
    base_query = str(query or "").strip()
    mode = str(chat_mode or "").strip().lower()
    context = _sanitize_playbook_context(playbook_context or {})
    action = _sanitize_guided_action(guided_action or {})
    if not mode and not context and not action:
        return base_query

    label = _CHAT_MODE_LABELS.get(mode, mode or "modo guiado")
    context_lines: List[str] = []
    if context.get("product"):
        context_lines.append(f"- produto_foco: {context['product']}")
    if context.get("segment"):
        context_lines.append(f"- segmento_foco: {context['segment']}")
    if context.get("une"):
        context_lines.append(f"- lojas_ou_une: {context['une']}")
    if context.get("period"):
        context_lines.append(f"- periodo: {context['period']}")
    if context.get("objective"):
        context_lines.append(f"- objetivo: {context['objective']}")

    action_lines: List[str] = []
    if action.get("actionLabel"):
        action_lines.append(f"- acao: {action['actionLabel']}")
    if action.get("source"):
        action_lines.append(f"- origem: {action['source']}")
    if action.get("playbookId"):
        action_lines.append(f"- playbook: {_CHAT_MODE_LABELS.get(str(action['playbookId']), str(action['playbookId']))}")
    if action.get("outputPreference"):
        action_lines.append(f"- formato_preferido: {action['outputPreference']}")
    if action.get("executionPolicy"):
        action_lines.append(f"- politica_execucao: {action['executionPolicy']}")
    if isinstance(action.get("directSend"), bool):
        action_lines.append(f"- envio_direto: {'sim' if action['directSend'] else 'nao'}")
    if action.get("toolHints"):
        action_lines.append(f"- tools_sugeridas: {', '.join(action['toolHints'])}")
    if action.get("missingDataBehavior"):
        action_lines.append(f"- politica_lacunas: {action['missingDataBehavior']}")

    sections: List[str] = [
        "Contexto operacional adicional para orientar esta resposta:",
        f"- modo analitico: {label}",
        "- preserve foco em decisao, acao e linguagem de negocio",
        "- use apenas dados reais observados no sistema",
        "- nao invente numeros, percentuais, precos ou fatos nao observados",
        "- quando houver calculo deterministico ou tool especifica, prefira esse caminho a uma resposta puramente narrativa",
        "- se faltar insumo critico para uma conclusao, diga exatamente o minimo que falta e continue com o que ja pode ser respondido",
    ]
    mode_instructions = _CHAT_MODE_INSTRUCTIONS.get(mode, [])
    if mode_instructions:
        sections.append("Diretrizes do modo:")
        sections.extend(f"- {item}" for item in mode_instructions)
    if context_lines:
        sections.append("Contexto informado:")
        sections.extend(context_lines)
    if action_lines:
        sections.append("Acao orientada:")
        sections.extend(action_lines)

    if not base_query:
        return "\n".join(sections).strip()
    return "\n".join([*sections, "Pergunta do usuario:", base_query]).strip()


def _extract_semantic_chat_query(query: str) -> str:
    text = str(query or "").strip()
    if not text:
        return ""

    for marker in ("Pergunta do usuario:", "Pergunta do usuário:", "[PERGUNTA_USUARIO]"):
        if marker in text:
            _, _, tail = text.partition(marker)
            cleaned = str(tail or "").strip()
            if cleaned:
                return cleaned

    return text


def _should_bypass_cache_for_query(query: str) -> bool:
    """
    Perguntas de mercado/concorrência usam dados externos voláteis e não devem
    reutilizar cache antigo por usuário.

    Pedidos estruturados de gráfico/tabela/dashboard/export também não devem
    reutilizar cache semântico fuzzy, porque pequenas mudanças de filtro
    (ex.: segmento, loja, produto) podem alterar completamente o payload.
    """
    q = (query or "").strip().lower()
    if _is_competitive_market_query(q):
        return True

    structured_output_markers = [
        "gráfico",
        "grafico",
        "dashboard",
        "painel",
        "tabela",
        "tabular",
        "export",
        "exportar",
        "planilha",
        "excel",
        "xlsx",
        "csv",
        "visualização",
        "visualizacao",
        "plot",
    ]
    if any(marker in q for marker in structured_output_markers):
        return True

    # Follow-ups contextuais devem recomputar resposta para preservar continuidade
    # da conversa e evitar repetição de payload anterior por cache semântico.
    contextual_markers = [
        "plano comercial",
        "plano de ação",
        "plano de acao",
        "com base nisso",
        "com base na última resposta",
        "com base na ultima resposta",
        "com base no relatório",
        "com base no relatorio",
        "da última resposta",
        "da ultima resposta",
        "continue",
        "detalhe",
        "próximas ações",
        "proximas acoes",
    ]
    if any(marker in q for marker in contextual_markers):
        return True

    short_followup_prefixes = ("e ", "agora ", "então ", "entao ")
    if len(q) <= 60 and any(q.startswith(prefix) for prefix in short_followup_prefixes):
        return True

    return False

# [OK] FIX: Import os for feedback endpoint
import os


def _augment_feedback_with_session_metadata(
    feedback_entry: Dict[str, Any],
    *,
    session_id: Optional[str],
    response_id: Optional[str],
    user_id: str,
) -> None:
    global session_manager

    if session_manager is None or not session_id or not response_id:
        return

    try:
        history_items = session_manager.get_full_history(session_id=session_id, user_id=user_id)
    except Exception as exc:
        logger.warning("feedback_session_metadata_lookup_failed: %s", exc)
        return

    for item in reversed(history_items):
        if str(item.get("role") or "") != "assistant":
            continue

        metadata = item.get("metadata") if isinstance(item.get("metadata"), dict) else {}
        if str(metadata.get("request_id") or item.get("id") or "") != str(response_id):
            continue

        citations = metadata.get("citations") if isinstance(metadata.get("citations"), list) else []
        if feedback_entry.get("source") in (None, "", []):
            feedback_entry["source"] = metadata.get("source")
        if feedback_entry.get("confidence") in (None, "", []):
            feedback_entry["confidence"] = metadata.get("confidence")
        if feedback_entry.get("mode") in (None, "", []):
            feedback_entry["mode"] = metadata.get("mode")
        if not feedback_entry.get("citations") and citations:
            feedback_entry["citations"] = citations
            feedback_entry["citations_count"] = len(citations)

        tool_names = metadata.get("tool_names")
        if isinstance(tool_names, list) and tool_names:
            feedback_entry["tool_names"] = tool_names
            feedback_entry["tool_call_count"] = int(metadata.get("tool_call_count") or len(tool_names))

        latency_seconds = metadata.get("latency_seconds")
        if isinstance(latency_seconds, (int, float)):
            feedback_entry["latency_seconds"] = float(latency_seconds)
            feedback_entry["latency_ms"] = round(float(latency_seconds) * 1000.0, 2)

        ab_variants = metadata.get("ab_variants")
        if isinstance(ab_variants, dict) and ab_variants:
            feedback_entry["ab_variants"] = {
                str(key): str(value)
                for key, value in ab_variants.items()
                if value not in (None, "", [])
            }
        break

async def initialize_agents_async():
    """
    Async initialization: Executed on app startup (background task).
    Ensures ChatService is ready when the user arrives.
    """
    global chat_service_v3, session_manager
    
    # Fast exit if already initialized
    if chat_service_v3 is not None:
        return

    async with _init_lock:
        if chat_service_v3 is not None:
            return
            
        logger.info("[START] [STARTUP] Initializing ChatServiceV3 (Metrics-First) in background...")
        
        try:
            # We must use asyncio.to_thread because initialization involves
            # heavy sync operations (loading vector stores, DuckDB connections)
            def _sync_init():
                logger.debug("ChatServiceV3 startup: entering synchronous initialization")
                try:
                    logger.debug("ChatServiceV3 startup: initializing SessionManager")
                    local_session_manager = SessionManager(
                        storage_dir=settings.SESSION_LEGACY_STORAGE_PATH,
                        db_path=settings.CHAT_STATE_DB_PATH,
                    )
                    
                    logger.debug("ChatServiceV3 startup: initializing ChatServiceV3")
                    local_service = ChatServiceV3(session_manager=local_session_manager)
                    
                    logger.info("ChatServiceV3 startup: initialization completed successfully")
                    return local_session_manager, local_service
                except Exception:
                    logger.exception("ChatServiceV3 startup: synchronous initialization failed")
                    raise
                    
            session_manager, chat_service_v3 = await asyncio.to_thread(_sync_init)
            
            logger.info("[OK] [STARTUP] ChatServiceV3 (Metrics-First) initialized successfully.")
        except Exception as e:
            logger.error(f"[ERROR] Failed to initialize ChatServiceV3: {e}", exc_info=True)
            # We don't raise here to avoid crashing the whole app, 
            # but Chat endpoints will fail if this didn't work.


def _persist_automation_state_to_history(automation_state: Dict[str, Any], current_user: User) -> None:
    if not isinstance(automation_state, dict):
        return
    if session_manager is None:
        return
    proposal_id = str(automation_state.get("proposal_id") or "").strip()
    session_id = str(automation_state.get("session_id") or "").strip()
    if not proposal_id or not session_id:
        return
    session_manager.update_message_metadata_by_request_id(
        session_id=session_id,
        user_id=str(getattr(current_user, "id", "") or ""),
        request_id=proposal_id,
        metadata_patch={
            "ui_payload": {
                "automation_request": automation_state,
            }
        },
    )


router = APIRouter(prefix="/chat", tags=["Chat"])


class ChatRequest(BaseModel):
    query: str
    session_id: Optional[str] = None
    chat_mode: Optional[str] = None
    playbook_context: Optional[Dict[str, Any]] = None
    guided_action: Optional[Dict[str, Any]] = None


class FeedbackRequest(BaseModel):
    response_id: str
    feedback_type: str
    comment: Optional[str] = None
    session_id: Optional[str] = None
    query_text: Optional[str] = None
    response_text: Optional[str] = None
    source: Optional[str] = None
    confidence: Optional[float] = None
    mode: Optional[str] = None
    citations: Optional[List[Dict[str, Any]]] = None


class AutomationProposalPayload(BaseModel):
    proposal_id: str
    action: str
    title: Optional[str] = None
    summary: Optional[str] = None
    request_text: Optional[str] = None
    session_id: Optional[str] = None
    params: Dict[str, Any] = Field(default_factory=dict)
    review_required: bool = False
    follow_up_action: Optional[str] = None
    follow_up_label: Optional[str] = None
    target_label: Optional[str] = None


class AutomationApproveRequest(BaseModel):
    approval_id: Optional[str] = None
    proposal: Optional[AutomationProposalPayload] = None
    follow_up_action: Optional[str] = None


class AutomationRejectRequest(BaseModel):
    approval_id: Optional[str] = None
    proposal: Optional[AutomationProposalPayload] = None


class ChatResponse(BaseModel):
    response: str


class ChatTableExportRequest(BaseModel):
    rows: List[Dict[str, Any]] = Field(default_factory=list)
    filename: Optional[str] = Field(default=None, max_length=120)
    caption: Optional[str] = Field(default=None, max_length=160)


@router.get("/market-research/download/{search_id}")
async def download_market_results(
    search_id: str,
    format: str = "xlsx"
):
    """
    Exporta os resultados da pesquisa de mercado para Excel ou CSV.
    """
    if search_id not in _market_search_cache:
        raise HTTPException(status_code=404, detail="Resultados da pesquisa não encontrados ou expirados.")

    data = _market_search_cache[search_id]
    rows = data["rows"]
    produto = data["produto"]
    internal = data.get("internal_price", {})

    output = io.BytesIO()
    
    if format.lower() == "xlsx":
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Pesquisa de Mercado"
        
        # Cabeçalho
        headers = ["Concorrente", "Produto", "Preço (R$)", "Fonte"]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
            
        # Dados
        for row in rows:
            ws.append([
                row["concorrente"],
                row["produto"],
                row["preco"],
                row["fonte"]
            ])
            
        # Adicionar Preço Interno se disponível
        if internal.get("preco_venda"):
            ws.append([])
            ws.append(["--- INFORMAÇÕES INTERNAS ---"])
            ws.append(["Nosso Preço de Venda", internal["preco_venda"]])
            ws.append(["Nosso Custo", internal.get("custo")])
            ws.append(["Mídia de Mercado", data.get("avg_price")])
            
        wb.save(output)
        output.seek(0)
        
        filename = f"pesquisa_mercado_{search_id}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        
    else: # Default to CSV
        import csv
        content = io.StringIO()
        writer = csv.writer(content, delimiter=";", lineterminator="\n")
        
        writer.writerow(["Concorrente", "Produto", "Preco", "Fonte"])
        for row in rows:
            writer.writerow([
                row["concorrente"],
                row["produto"],
                f"{row['preco']:.2f}".replace(".", ","),
                row["fonte"]
            ])
            
        if internal.get("preco_venda"):
            writer.writerow([])
            writer.writerow(["--- INFORMACOES INTERNAS ---"])
            writer.writerow(["Nosso Preco de Venda", f"{internal['preco_venda']:.2f}".replace(".", ",")])
            writer.writerow(["Nosso Custo", f"{internal.get('custo', 0):.2f}".replace(".", ",")])
            
        output.write(content.getvalue().encode("utf-8-sig"))
        output.seek(0)
        
        filename = f"pesquisa_mercado_{search_id}.csv"
        media_type = "text/csv"

    return StreamingResponse(
        output,
        media_type=media_type,
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


def _sanitize_chat_export_filename(raw_value: Optional[str], fallback: str = "chat_tabela") -> str:
    normalized = re.sub(r"[^A-Za-z0-9_-]+", "_", str(raw_value or "").strip()).strip("._-")
    return (normalized[:80] or fallback).lower()


def _coerce_chat_export_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return value


@router.post("/table-export")
async def export_chat_table(
    payload: ChatTableExportRequest,
    format: str = "xlsx",
    current_user: Annotated[User, Depends(get_current_active_user)] = None,
):
    rows = [row for row in payload.rows if isinstance(row, dict)]
    if not rows:
        raise HTTPException(status_code=400, detail="Nenhuma linha válida foi informada para exportação.")

    headers: List[str] = []
    for row in rows:
        for key in row.keys():
            label = str(key or "").strip()
            if label and label not in headers:
                headers.append(label)

    if not headers:
        raise HTTPException(status_code=400, detail="A tabela não possui colunas válidas para exportação.")

    filename_base = _sanitize_chat_export_filename(payload.filename or payload.caption or "chat_tabela")
    output = io.BytesIO()
    normalized_format = str(format or "xlsx").strip().lower()

    if normalized_format == "xlsx":
        import openpyxl
        from openpyxl.styles import Alignment, Font

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = (payload.caption or "Tabela chat")[:31]

        worksheet.append(headers)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        for row in rows:
            worksheet.append([_coerce_chat_export_cell(row.get(header)) for header in headers])

        workbook.save(output)
        output.seek(0)
        filename = f"{filename_base}.xlsx"
        media_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    elif normalized_format == "csv":
        import csv

        content = io.StringIO()
        writer = csv.writer(content, delimiter=";", lineterminator="\n")
        writer.writerow(headers)
        for row in rows:
            writer.writerow([_coerce_chat_export_cell(row.get(header)) for header in headers])

        output.write(content.getvalue().encode("utf-8-sig"))
        output.seek(0)
        filename = f"{filename_base}.csv"
        media_type = "text/csv"
    else:
        raise HTTPException(status_code=400, detail="Formato de exportação inválido. Use csv ou xlsx.")

    return StreamingResponse(
        output,
        media_type=media_type,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@router.post("/stream-token")
async def create_stream_token(
    request: Request,
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    """
    Emite token efêmero para SSE, evitando expor JWT completo na URL.
    TTL curto e uso único.
    """
    if not _is_chat_allowed_for_user(current_user):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="ChatBI em canary fechado para este perfil. Solicite liberacao do acesso.",
        )

    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        raise HTTPException(status_code=401, detail="Authorization Bearer ausente")

    bearer = auth_header[7:].strip()
    if not bearer:
        raise HTTPException(status_code=401, detail="Token inválido")

    ephemeral = issue_stream_token(bearer)
    return {"stream_token": ephemeral, "expires_in": 120}


@router.get("/llm/status")
async def llm_status(
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    """
    Status operacional dos provedores LLM configurados para o ChatBI.
    """
    global chat_service_v3
    if chat_service_v3 is None:
        await initialize_agents_async()
    if chat_service_v3 is None:
        raise HTTPException(status_code=503, detail="Chat service ainda não inicializado")
    return chat_service_v3.get_llm_status()


@router.get("/context7/status")
async def context7_status(
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    """
    Status operacional da integração Context7 externa (quando configurada).
    """
    from backend.app.core.integrations.context7_status import get_context7_status

    return get_context7_status()


@router.get("/report-templates")
async def list_official_chat_report_templates(
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    """
    Catalogo oficial de templates de relatorio do ChatBI (Fase 3).
    """
    return {
        "templates": get_official_report_templates(),
        "chat_canary_enabled": settings.CHAT_CANARY_ENABLED,
    }


@router.get("/stream")
async def stream_chat(
    q: str,
    session_id: str,
    request: Request,
    token: Annotated[str, Depends(get_token_from_header_or_query)],
    chat_mode: Optional[str] = None,
    playbook_context: Optional[str] = None,
    guided_action: Optional[str] = None,
):
    """
    Streaming endpoint using Server-Sent Events (SSE)
    Integrates the agent system for dynamic responses.
    """
    from backend.app.api.dependencies import get_current_user_from_token
    from backend.app.core.context import set_current_user_context

    try:
        current_user = await get_current_user_from_token(token)
        # [OK] CRITICAL FIX: Set context for tools running in background
        set_current_user_context(current_user)
        logger.info(f"SSE authenticated user: {current_user.username}")
        if not _is_chat_allowed_for_user(current_user):
            return JSONResponse(
                status_code=status.HTTP_403_FORBIDDEN,
                content={"detail": "ChatBI em canary fechado para este perfil. Solicite liberacao do acesso."},
            )
    except Exception as e:
        logger.error(f"SSE authentication failed: {e}")
        return JSONResponse(
            status_code=status.HTTP_401_UNAUTHORIZED,
            content={"detail": "Sessao invalida ou expirada. Faca login novamente."},
        )

    parsed_playbook_context = _parse_playbook_context_param(playbook_context)
    parsed_guided_action = _parse_guided_action_param(guided_action)
    effective_query = _build_guided_chat_query(q, chat_mode, parsed_playbook_context, parsed_guided_action)
    last_event_id = request.headers.get("Last-Event-ID")
    logger.info(f"==> SSE STREAM REQUEST: {q} (Session: {session_id}) (Last-Event-ID: {last_event_id}) <==")
    stream_request_id = str(_uuid4())
    user_capabilities = get_chat_capabilities_for_user(current_user)
    trace_logger.info(
        "chat_sse_stream_started",
        request_id=stream_request_id,
        session_id=str(session_id),
        user_id=str(getattr(current_user, "id", "")),
        role=str(getattr(current_user, "role", "")),
        query_excerpt=str(q)[:160],
        guided_mode=str(chat_mode or ""),
        guided_context=parsed_playbook_context,
        guided_action=parsed_guided_action,
        last_event_id=last_event_id,
    )

    async def event_generator():
        final_sent = False
        response_request_id = stream_request_id
        final_event_payload: Dict[str, Any] = {
            "type": "final",
            "text": "",
            "done": True,
            "request_id": response_request_id,
        }
        try:
            event_counter = int(last_event_id) if last_event_id else 0

            def _update_final_event_metadata(payload: Optional[Dict[str, Any]]) -> None:
                nonlocal response_request_id

                if not isinstance(payload, dict):
                    return

                response_request_id = str(payload.get("request_id") or response_request_id)
                final_event_payload["request_id"] = response_request_id

                for key in (
                    "source",
                    "confidence",
                    "mode",
                    "image_asset",
                    "audio_asset",
                    "automation_request",
                    "chart_data",
                    "chart_spec",
                    "table_data",
                    "dashboard_spec",
                ):
                    value = payload.get(key)
                    if value not in (None, ""):
                        final_event_payload[key] = value

                if "chart_data" in final_event_payload and "chart_spec" not in final_event_payload:
                    final_event_payload["chart_spec"] = final_event_payload["chart_data"]

                citations_value = sanitize_citations(payload.get("citations"))
                if citations_value:
                    final_event_payload["citations"] = citations_value

            async def _persist_session_turn(
                response_payload: Dict[str, Any],
                metadata_query: Optional[str] = None,
            ) -> None:
                try:
                    global chat_service_v3
                    if chat_service_v3 is None:
                        await initialize_agents_async()
                    if chat_service_v3 is None:
                        return
                    if not user_capabilities.get("memory", False):
                        return

                    effective_query = str(metadata_query or q)
                    user_metadata = chat_service_v3.build_session_message_metadata(query=effective_query, role="user")
                    chat_service_v3.session_manager.add_message(
                        session_id,
                        "user",
                        q,
                        current_user.id,
                        metadata=user_metadata,
                    )

                    response_text_local = ""
                    result_payload = response_payload.get("result", {})
                    if isinstance(result_payload, dict):
                        response_text_local = str(result_payload.get("mensagem", ""))
                    if not response_text_local:
                        response_text_local = str(response_payload.get("response") or "")

                    assistant_metadata = chat_service_v3.build_session_message_metadata(
                        query=effective_query,
                        response=response_payload,
                        role="assistant",
                    )
                    chat_service_v3.session_manager.add_message(
                        session_id,
                        "assistant",
                        response_text_local,
                        current_user.id,
                        metadata=assistant_metadata,
                    )
                except Exception as persist_error:
                    logger.warning(f"Falha ao persistir histórico do fast path SSE: {persist_error}")

            # --- FAST PATH: Detecção de Saudações Simples (Zero Latency) ---
            # EXTREMAMENTE CRÍTICO: Deve rodar ANTES de qualquer inicialização pesada (initialize_agents_async)
            query_clean = q.strip().lower()
            greetings = [
                "oi", "ola", "olá", "bom dia", "boa tarde", "boa noite",
                "hello", "hi", "eai", "opa", "teste", "funcionando?", "tá funcionando", "ta funcionando",
                "tudo bem", "como vai", "e aí"
            ]

            # [OK] FIX 2026-01-14: Aumentado limite de 20 para 40 caracteres
            # "ola boa tarde tudo bem" tem 21 chars e não entrava no fast path
            # [OK] FIX 2026-01-17: Lógica refinada para não interceptar queries reais
            # Só interceptar se for APENAS saudação ou saudação curta SEM verbos de ação
            action_keywords = [
                "analis", "venda", "estoque", "ruptura", "grafico", "relatorio", 
                "quanto", "quais", "mostre", "gere", "crie", "veja", "dados"
            ]
            has_action = any(k in query_clean for k in action_keywords)
            
            is_pure_greeting = query_clean in greetings
            # Reduzido de 40 para 20 caracteres para evitar falsos positivos
            is_short_greeting = len(query_clean) < 20 and any(g in query_clean for g in greetings)

            if (is_pure_greeting or is_short_greeting) and not has_action:
                trace_logger.info(
                    "chat_sse_fast_path_selected",
                    request_id=response_request_id,
                    session_id=str(session_id),
                    user_id=str(getattr(current_user, "id", "")),
                    fast_path="greeting",
                )
                import random
                import asyncio
                from datetime import datetime

                # [OK] FIX 2026-01-15: Responder com saudação apropriada ao período
                # Detecta período mencionado pelo usuário ou usa hora atual
                if "boa noite" in query_clean:
                    saudacao = "Boa noite"
                elif "boa tarde" in query_clean:
                    saudacao = "Boa tarde"
                elif "bom dia" in query_clean:
                    saudacao = "Bom dia"
                else:
                    # Usa hora atual do servidor
                    hora = datetime.now().hour
                    if 5 <= hora < 12:
                        saudacao = "Bom dia"
                    elif 12 <= hora < 18:
                        saudacao = "Boa tarde"
                    else:
                        saudacao = "Boa noite"

                responses = [
                    f"{saudacao}! Sou seu assistente de BI. Como posso ajudar com os dados hoje?",
                    f"{saudacao}! Tudo pronto para analisar seus dados. O que você gostaria de ver?",
                    f"{saudacao}! Estou à disposição. Pode me pedir gráficos, relatórios ou análises.",
                    f"{saudacao}! Vamos encontrar alguns insights? É só perguntar."
                ]
                response_text = random.choice(responses)
                
                # Simular steps de progresso para UX consistente
                event_counter += 1
                yield f"id: {event_counter}\n"
                yield f"data: {safe_json_dumps({'type': 'tool_progress', 'tool': 'system.thinking', 'status': 'start'})}\n\n"
                
                await asyncio.sleep(0.1) 
                
                event_counter += 1
                yield f"id: {event_counter}\n"
                yield f"data: {safe_json_dumps({'type': 'tool_progress', 'tool': 'system.finalizing', 'status': 'finishing'})}\n\n"

                # Stream response text
                words = response_text.split(" ")
                for i in range(0, len(words), 3):
                    chunk_words = words[i:i + 3]
                    prefix = " " if i > 0 else ""
                    chunk_text = prefix + " ".join(chunk_words)
                    event_counter += 1
                    yield f"id: {event_counter}\n"
                    yield f"data: {safe_json_dumps({'type': 'text', 'text': chunk_text, 'done': False})}\n\n"
                    await asyncio.sleep(0.05) # Typing effect

                # Finalize
                event_counter += 1
                yield f"id: {event_counter}\n"
                yield f"data: {safe_json_dumps({'type': 'final', 'text': '', 'done': True, 'request_id': response_request_id})}\n\n"
                final_sent = True
                await _persist_session_turn(
                    {
                        "type": "text",
                        "result": {"mensagem": response_text},
                    },
                    metadata_query=effective_query,
                )
                return # SAÍDA ANTECIPADA - Evita carregar o agente pesado

            # --- FAST PATH: pesquisa concorrencial/mercado direta (sem pipeline completo do agente) ---
            if _is_competitive_market_query(query_clean):
                trace_logger.info(
                    "chat_sse_fast_path_selected",
                    request_id=response_request_id,
                    session_id=str(session_id),
                    user_id=str(getattr(current_user, "id", "")),
                    fast_path="market_research",
                )
                session_history = []
                try:
                    if session_manager is not None:
                        session_history = session_manager.get_history(session_id, current_user.id)
                except Exception as history_error:
                    logger.warning(f"Falha ao obter histórico para fast path de mercado: {history_error}")

                resolved_market_query = _resolve_competitive_market_followup_query(q, session_history)
                resolved_market_clean = resolved_market_query.strip().lower()

                # Pesquisa de mercado genérica usa pesquisa aberta multi-provider.
                # Pesquisa concorrencial fica reservada para concorrentes explícitos.
                use_market_web = _should_use_market_web_fast_path(resolved_market_clean)
                tool_label = 'tool.market_research' if use_market_web else 'tool.competitive_research'
                default_source = "tool.pesquisar_mercado_web" if use_market_web else "tool.pesquisar_precos_concorrentes"

                event_counter += 1
                yield f"id: {event_counter}\n"
                yield f"data: {safe_json_dumps({'type': 'tool_progress', 'tool': tool_label, 'status': 'start'})}\n\n"

                if use_market_web:
                    fast_path_result = await _run_market_research_fast_path(resolved_market_query, return_payload=True)
                else:
                    fast_path_result = await _run_competitive_market_fast_path(resolved_market_query, return_payload=True)
                payload = fast_path_result.get("payload", {}) if isinstance(fast_path_result, dict) else {}
                market_contract = _market_contract_from_payload(payload, default_source=default_source)
                response_text = str(fast_path_result.get("text") or "") if isinstance(fast_path_result, dict) else str(fast_path_result or "")
                response_text = _sanitize_response_for_role(response_text, getattr(current_user, "role", "user"))

                event_counter += 1
                yield f"id: {event_counter}\n"
                yield f"data: {safe_json_dumps({'type': 'tool_progress', 'tool': 'system.finalizing', 'status': 'finishing'})}\n\n"

                words = response_text.split(" ")
                for i in range(0, len(words), 10):
                    chunk_words = words[i:i + 10]
                    prefix = " " if i > 0 else ""
                    chunk_text = prefix + " ".join(chunk_words)
                    event_counter += 1
                    yield f"id: {event_counter}\n"
                    yield f"data: {safe_json_dumps({'type': 'text', 'text': chunk_text, 'done': False})}\n\n"

                event_counter += 1
                yield f"id: {event_counter}\n"
                yield f"data: {safe_json_dumps({'type': 'final', 'text': '', 'done': True, 'request_id': response_request_id, **market_contract})}\n\n"
                final_sent = True
                await _persist_session_turn(
                    {
                        "type": "text",
                        "result": {"mensagem": response_text},
                        **market_contract,
                    },
                    metadata_query=_build_guided_chat_query(
                        resolved_market_query,
                        chat_mode,
                        parsed_playbook_context,
                        parsed_guided_action,
                    ),
                )
                return

            # --- FAST PATH: perguntas determinísticas de KPI (sem LLM) ---
            kpi_intents = [
                "kpi",
                "kpis",
                "indicadores",
                "metricas",
                "métricas",
                "resumo executivo",
            ]
            if any(term in query_clean for term in kpi_intents) and len(query_clean) <= 60:
                trace_logger.info(
                    "chat_sse_fast_path_selected",
                    request_id=response_request_id,
                    session_id=str(session_id),
                    user_id=str(getattr(current_user, "id", "")),
                    fast_path="kpi_hint",
                )
                deterministic_msg = (
                    "Para KPIs instantâneos, use o Dashboard/endpoint de métricas. "
                    "Posso detalhar um KPI específico se você informar qual (ex.: venda_30dd, margem, estoque)."
                )
                event_counter += 1
                yield f"id: {event_counter}\n"
                yield f"data: {safe_json_dumps({'type': 'text', 'text': deterministic_msg, 'done': False})}\n\n"
                event_counter += 1
                yield f"id: {event_counter}\n"
                yield f"data: {safe_json_dumps({'type': 'final', 'text': '', 'done': True, 'request_id': response_request_id})}\n\n"
                final_sent = True
                await _persist_session_turn(
                    {
                        "type": "text",
                        "result": {"mensagem": deterministic_msg},
                    },
                    metadata_query=effective_query,
                )
                return
            # ---------------------------------------------------------------------

            # [DEBUG] FIX: Ensure initialization if startup task hasn't finished yet
            if chat_service_v3 is None:
                logger.info("[RETRY] Agent system not ready yet. Waiting for initialization...")
                await initialize_agents_async()

            if chat_service_v3 is None:
                trace_logger.error(
                    "chat_sse_stream_failed",
                    request_id=response_request_id,
                    session_id=str(session_id),
                    user_id=str(getattr(current_user, "id", "")),
                    error="agent_system_unavailable",
                )
                yield f"data: {safe_json_dumps({'type': 'error', 'error': 'Agent system could not be initialized', 'request_id': response_request_id})}\n\n"
                return

            # Retrieve History - Corrected with user_id for security
            # chat_history = session_manager.get_history(session_id, current_user.id)
            # Add User Message to History immediately
            # session_manager.add_message(session_id, "user", q, current_user.id)

            logger.info(f"Processing query with ChatServiceV3: '{q}'")

            # [OK] FIX 2026-01-14: Cache agora usa user_id para isolamento
            # Isso evita que dados de UNE 1685 sejam retornados para query de UNE 1700
            user_cache_id = str(current_user.id) if current_user else None
            metrics = MetricsService()

            # Cache policy: consultas de mercado/concorrência não reutilizam cache.
            bypass_cache = _should_bypass_cache_for_query(effective_query)

            # NOVO: Verificar Semantic Cache primeiro (com user_id)
            cache_key_query = _build_chat_cache_key(session_id, effective_query)
            metrics.increment("chat_cache_lookups_total")
            cached_response = None if bypass_cache else cache_get(cache_key_query, user_id=user_cache_id)
            if cached_response and not _is_degraded_or_error_response(cached_response):
                metrics.increment("chat_cache_hits_total")
                logger.info(f"CACHE HIT: Resposta encontrada em cache para: {q[:50]}... (user={user_cache_id})")
                trace_logger.info(
                    "chat_sse_cache_hit",
                    request_id=response_request_id,
                    session_id=str(session_id),
                    user_id=str(getattr(current_user, "id", "")),
                    cache_key=cache_key_query,
                    bypass_cache=bool(bypass_cache),
                )
                # Mesmo em cache-hit, manter histórico consistente para follow-ups.
                try:
                    if chat_service_v3 is not None:
                        user_metadata = chat_service_v3.build_session_message_metadata(
                            query=effective_query,
                            role="user",
                        )
                        chat_service_v3.session_manager.add_message(
                            session_id,
                            "user",
                            q,
                            current_user.id,
                            metadata=user_metadata,
                        )
                        cached_text = ""
                        if isinstance(cached_response, dict):
                            res = cached_response.get("result", {})
                            if isinstance(res, dict):
                                cached_text = str(res.get("mensagem", ""))
                            else:
                                cached_text = str(res)
                        if not cached_text:
                            cached_text = str(cached_response)
                        assistant_metadata = chat_service_v3.build_session_message_metadata(
                            query=effective_query,
                            response=cached_response if isinstance(cached_response, dict) else None,
                            role="assistant",
                        )
                        chat_service_v3.session_manager.add_message(
                            session_id,
                            "assistant",
                            cached_text,
                            current_user.id,
                            metadata=assistant_metadata,
                        )
                except Exception as e:
                    logger.warning(f"Falha ao registrar histórico em cache-hit: {e}")
                event_counter += 1
                yield f"id: {event_counter}\n"
                yield f"data: {safe_json_dumps({'type': 'cache_hit', 'done': False})}\n\n"
                agent_response = cached_response
            elif cached_response:
                metrics.increment("chat_cache_misses_total")
                logger.info("CACHE SKIP: resposta degradada/erro não será reutilizada")
                trace_logger.info(
                    "chat_sse_cache_skip",
                    request_id=response_request_id,
                    session_id=str(session_id),
                    user_id=str(getattr(current_user, "id", "")),
                    cache_key=cache_key_query,
                    reason="degraded_or_error_response",
                )
                agent_response = None
            else:
                metrics.increment("chat_cache_misses_total")
                trace_logger.info(
                    "chat_sse_cache_miss",
                    request_id=response_request_id,
                    session_id=str(session_id),
                    user_id=str(getattr(current_user, "id", "")),
                    cache_key=cache_key_query,
                    bypass_cache=bool(bypass_cache),
                )
                # OPTIMIZATION 2025: Stream progress events during agent execution
                import asyncio
                event_queue = asyncio.Queue()

                async def progress_callback(event):
                    await event_queue.put(event)

                # [OK] FIX: Timeout reduzido de 300s para 60s (resposta mais rápida)
                trace_logger.info(
                    "chat_async_job_started",
                    request_id=response_request_id,
                    session_id=str(session_id),
                    user_id=str(getattr(current_user, "id", "")),
                    job_name="chat_service_process_message",
                )
                agent_task = asyncio.create_task(
                    asyncio.wait_for(
                        chat_service_v3.process_message(
                            query=effective_query,
                            session_id=session_id, 
                            user_id=current_user.id,
                            user_role=current_user.role,
                            user_capabilities=user_capabilities,
                            request_id=response_request_id,
                            on_progress=progress_callback
                        ),
                        timeout=90.0  # [OK] FIX: Aumentado para 90s para queries complexas com gráficos
                    )
                )

                # Stream progress events as they arrive
                agent_response = None
                keepalive_counter = 0
                keepalive_interval = _SSE_KEEPALIVE_INTERVAL_TICKS

                while True:
                    try:
                        event = await asyncio.wait_for(event_queue.get(), timeout=_SSE_EVENT_POLL_TIMEOUT_SECONDS)
                        event_counter += 1
                        keepalive_counter = 0  # Reset keepalive on real event
                        yield f"id: {event_counter}\n"
                        yield f"data: {safe_json_dumps(event)}\n\n"
                    except asyncio.TimeoutError:
                        keepalive_counter += 1

                        # Send keepalive event every 5 seconds to prevent frontend timeout
                        if keepalive_counter >= keepalive_interval:
                            event_counter += 1
                            keepalive_event = {"type": "keepalive", "message": "Ainda processando sua análise complexa..."}
                            yield f"id: {event_counter}\n"
                            yield f"data: {safe_json_dumps(keepalive_event)}\n\n"
                            keepalive_counter = 0

                        if agent_task.done():
                            try:
                                agent_response = agent_task.result()
                                trace_logger.info(
                                    "chat_async_job_completed",
                                    request_id=response_request_id,
                                    session_id=str(session_id),
                                    user_id=str(getattr(current_user, "id", "")),
                                    job_name="chat_service_process_message",
                                )
                            except asyncio.TimeoutError:
                                logger.error(f"Agent timeout após 90s para query: {q}")
                                trace_logger.error(
                                    "chat_async_job_failed",
                                    request_id=response_request_id,
                                    session_id=str(session_id),
                                    user_id=str(getattr(current_user, "id", "")),
                                    job_name="chat_service_process_message",
                                    error="timeout",
                                )
                                if _is_competitive_market_query(q):
                                    recovered_result = await _run_competitive_market_fast_path(q, return_payload=True)
                                    recovered_payload = recovered_result.get("payload", {}) if isinstance(recovered_result, dict) else {}
                                    recovered_text = (
                                        str(recovered_result.get("text") or "")
                                        if isinstance(recovered_result, dict)
                                        else str(recovered_result or "")
                                    )
                                    agent_response = {
                                        "type": "text",
                                        "result": {"mensagem": recovered_text},
                                        "source": recovered_payload.get("source"),
                                        "confidence": recovered_payload.get("confidence"),
                                        "mode": recovered_payload.get("mode"),
                                        "citations": recovered_payload.get("citations"),
                                    }
                                else:
                                    agent_response = {
                                        "type": "text",
                                        "result": {
                                            "mensagem": "O tempo limite de processamento foi excedido (90 segundos). Tente uma pergunta mais objetiva para receber a resposta mais rápido."
                                        }
                                    }
                            except Exception as e:
                                logger.error(f"Agent error: {e}", exc_info=True)
                                trace_logger.error(
                                    "chat_async_job_failed",
                                    request_id=response_request_id,
                                    session_id=str(session_id),
                                    user_id=str(getattr(current_user, "id", "")),
                                    job_name="chat_service_process_message",
                                    error=str(e),
                                )
                                agent_response = {
                                    "type": "text",
                                    "result": {
                                        "mensagem": "Nao foi possivel concluir a analise agora. Tente novamente em instantes."
                                    }
                                }
                            break

                # [OK] FIX 2026-01-14: Salvar resposta válida em cache COM user_id
                if (
                    (not bypass_cache)
                    and agent_response
                    and "error" not in str(agent_response).lower()
                    and not _is_degraded_or_error_response(agent_response)
                ):
                    cache_set(cache_key_query, agent_response, user_id=user_cache_id)
            
            if not agent_response:
                logger.warning(f"Agent retornou resposta vazia para query: {q}")
                agent_response = {
                    "type": "text",
                    "result": {
                        "mensagem": "Não foi possível concluir a análise agora. Verifique limites de quota/billing e tente uma pergunta mais objetiva."
                    }
                }
            
            if isinstance(agent_response, dict):
                _update_final_event_metadata(agent_response)

            semantic_query = _extract_semantic_chat_query(effective_query)
            validation_context = (
                _build_stream_validation_context(semantic_query, agent_response)
                if isinstance(agent_response, dict)
                else {}
            )
            # Validação de qualidade da resposta (guardrail enterprise)
            validation = validate_response(agent_response, semantic_query, context=validation_context)
            if isinstance(agent_response, dict) and getattr(validation, "should_block", False):
                trace_logger.warning(
                    "chat_stream_response_blocked_by_validator",
                    request_id=response_request_id,
                    session_id=str(session_id),
                    user_id=str(getattr(current_user, "id", "")),
                    expected_capability=validation_context.get("expected_capability"),
                    actual_capability=validation_context.get("actual_capability"),
                    block_reason=getattr(validation, "block_reason", None),
                    issues=getattr(validation, "issues", []),
                )
                agent_response = _build_stream_validation_block_response(
                    query=semantic_query,
                    validation_result=validation,
                    validation_context=validation_context,
                )
                _update_final_event_metadata(agent_response)

            response_type = agent_response.get("type", "text")
            response_content = agent_response.get("result")
            response_text = ""

            if response_type in {"text", "tool_result", "dashboard"}:
                # CRITICAL FIX: Check if tool_result contains chart_data from chart generation tools
                result_data = agent_response.get("result", {})
                if not isinstance(result_data, dict):
                    result_data = {}

                dashboard_spec = agent_response.get("dashboard_spec") or result_data.get("dashboard_spec")
                if isinstance(dashboard_spec, str):
                    try:
                        dashboard_spec = json.loads(dashboard_spec)
                    except json.JSONDecodeError:
                        logger.error("Failed to parse dashboard_spec JSON string")
                        dashboard_spec = None

                if isinstance(dashboard_spec, dict):
                    event_counter += 1
                    yield f"id: {event_counter}\n"
                    yield "data: " + safe_json_dumps(
                        {
                            "type": "dashboard",
                            "dashboard_spec": dashboard_spec,
                            "request_id": response_request_id,
                            "done": False,
                        }
                    ) + "\n\n"

                # [OK] FIX 2026-01-17: Check top-level chart_data FIRST (ChatServiceV3 format)
                chart_data = agent_response.get("chart_data")

                # Fallback: Check inside result dict (legacy format)
                if not chart_data:
                    chart_data = result_data.get("chart_data") or result_data.get("chart_spec")

                # [OK] STREAM CHART IF FOUND (either top-level or legacy)
                if chart_data and not isinstance(dashboard_spec, dict):
                    logger.info("Chart data detected - streaming chart to frontend")
                    # Parse chart_data if it's a JSON string
                    if isinstance(chart_data, str):
                        try:
                            chart_data = json.loads(chart_data)
                        except json.JSONDecodeError:
                            logger.error("Failed to parse chart_data JSON string")
                            chart_data = None

                    if chart_data:
                        # Stream the chart
                        event_counter += 1
                        yield f"id: {event_counter}\n"
                        yield f"data: {safe_json_dumps({'type': 'chart', 'chart_spec': chart_data, 'request_id': response_request_id, 'done': False})}\n\n"

                table_data = agent_response.get("table_data") or result_data.get("table_data")
                if isinstance(table_data, list) and table_data:
                    event_counter += 1
                    yield f"id: {event_counter}\n"
                    yield f"data: {safe_json_dumps({'type': 'table', 'data': table_data, 'request_id': response_request_id, 'done': False})}\n\n"

                response_text = result_data.get("mensagem", "")

                if not response_text or (isinstance(response_text, str) and not response_text.strip()):
                    if isinstance(chart_data, dict):
                        response_text = "Gráfico gerado com sucesso."
                    elif isinstance(dashboard_spec, dict):
                        response_text = "Dashboard gerado com sucesso."
                    else:
                        response_text = "Resposta processada."

                if not isinstance(response_text, str):
                    response_text = str(response_text)
            
            
            elif response_type == "code_result":
                # Lógica V2: O LangGraph abstrai "code_result" para "tool_result" ou texto direto
                # Mas mantemos compatibilidade caso o output seja complexo
                chart_spec = agent_response.get("chart_spec")
                response_text = agent_response.get("text_override") or str(response_content)

                if chart_spec:
                    event_counter += 1
                    yield f"id: {event_counter}\n"
                    yield f"data: {safe_json_dumps({'type': 'chart', 'chart_spec': chart_spec, 'request_id': response_request_id, 'done': False})}\n\n"
            
            # Reforço de qualidade interno: não expor detalhes técnicos ao usuário final.
            if response_text and response_text.strip() and not validation.is_valid:
                logger.warning(
                    f"[QUALITY] Resposta com baixa confiança para query='{q[:120]}': "
                    f"issues={validation.issues[:3]} suggestions={validation.suggestions[:2]}"
                )

            if response_text and response_text.strip():
                sanitized = _sanitize_response_for_role(response_text, getattr(current_user, "role", "user"))
                if sanitized != response_text:
                    logger.info("[QUALITY] Sanitizacao de saida aplicada para remover termos tecnicos.")
                response_text = sanitized

            # Só fazer streaming de texto se houver texto para enviar
            if response_text and response_text.strip():
                words = response_text.split(" ")
                # [OK] FIX: Otimizado de 1 para 8 palavras por chunk (8x mais rápido)
                chunk_size = 8 
                
                logger.info(f"Initiating text streaming of {len(words)} words...")
                
                for i in range(0, len(words), chunk_size):
                    chunk_words = words[i:i + chunk_size]
                    # Reconstruct spacing correctly
                    prefix = " " if i > 0 else ""
                    chunk_text = prefix + " ".join(chunk_words)
                    
                    event_counter += 1

                    yield f"id: {event_counter}\n"
                    yield f"data: {safe_json_dumps({'type': 'text', 'text': chunk_text, 'done': False})}\n\n"
                    
                    # Small delay to simulate typing speed if needed, but usually network latency is enough
                    # await asyncio.sleep(0.01)

            logger.info("Text streaming complete. Sending done signal.")

        except APIError as e:
            logger.error(f"Agent API Error in stream: {e.message}", exc_info=True)
            trace_logger.error(
                "chat_sse_stream_failed",
                request_id=response_request_id,
                session_id=str(session_id),
                user_id=str(getattr(current_user, "id", "")),
                error=e.message,
            )
            yield f"data: {safe_json_dumps({'type': 'error', 'error': e.message, 'details': e.details, 'request_id': response_request_id})}\n\n"
        except Exception as e:
            error_msg = str(e)
            logger.error(f"Unexpected error in stream: {error_msg}", exc_info=True)
            trace_logger.error(
                "chat_sse_stream_failed",
                request_id=response_request_id,
                session_id=str(session_id),
                user_id=str(getattr(current_user, "id", "")),
                error=error_msg,
            )

            # Generic user-friendly error (never expose technical details)
            error_response = {
                'type': 'error',
                'error': 'Não foi possível processar sua solicitação no momento. Por favor, tente novamente.',
                'error_type': 'generic',
                'request_id': response_request_id,
            }

            yield f"data: {safe_json_dumps(error_response)}\n\n"
        finally:
            # 🛑 SAFETY NET: Always send DONE signal to prevent frontend infinite spinner
            if not final_sent:
                yield f"data: {safe_json_dumps(final_event_payload)}\n\n"
            trace_logger.info(
                "chat_sse_stream_finished",
                request_id=response_request_id,
                session_id=str(session_id),
                user_id=str(getattr(current_user, "id", "")),
                final_sent=bool(final_sent),
                total_events=event_counter,
                source=final_event_payload.get("source"),
                mode=final_event_payload.get("mode"),
            )

    
    return StreamingResponse(
        event_generator(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "Connection": "keep-alive",
            "X-Accel-Buffering": "no",
        }
    )


@router.post("/feedback")
async def submit_feedback(
    feedback_data: FeedbackRequest,
    current_user: Annotated[User, Depends(get_current_active_user)],
):
    citations = sanitize_citations(feedback_data.citations)
    feedback_entry = {
        "timestamp": datetime.now().isoformat(),
        "user_id": str(getattr(current_user, "username", None) or getattr(current_user, "id", "anonymous")),
        "user_uuid": str(getattr(current_user, "id", "anonymous")),
        "request_id": feedback_data.response_id,
        "response_id": feedback_data.response_id,
        "feedback_type": feedback_data.feedback_type,
        "comment": feedback_data.comment,
        "session_id": feedback_data.session_id,
        "query_text": feedback_data.query_text,
        "response_text": feedback_data.response_text,
        "source": feedback_data.source,
        "confidence": feedback_data.confidence,
        "mode": feedback_data.mode,
        "citations": citations,
        "citations_count": len(citations),
    }
    _augment_feedback_with_session_metadata(
        feedback_entry,
        session_id=feedback_data.session_id,
        response_id=feedback_data.response_id,
        user_id=str(getattr(current_user, "id", "anonymous")),
    )
    
    feedback_file_path = Path(settings.LEARNING_FEEDBACK_PATH) / "feedback.jsonl"
    os.makedirs(Path(settings.LEARNING_FEEDBACK_PATH), exist_ok=True)
    try:
        with open(feedback_file_path, "a", encoding="utf-8") as f:
            f.write(safe_json_dumps(feedback_entry, ensure_ascii=False) + "\n")
        logger.info(
            "Feedback submitted by %s: %s",
            str(getattr(current_user, "username", None) or getattr(current_user, "id", "anonymous")),
            feedback_entry,
        )
    except OSError as e:
        logger.error(f"Failed to write feedback to file: {e}", exc_info=True)
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Nao foi possivel salvar o feedback agora."
        )

    chat_state_feedback_status = "skipped"
    try:
        memory_agent = get_memory_agent()
        rating_map = {"positive": 5, "partial": 3, "negative": 1}
        await memory_agent.save_feedback(
            request_id=feedback_data.response_id,
            rating=rating_map.get(str(feedback_data.feedback_type or "").strip().lower(), 3),
            comment=feedback_data.comment,
        )
        chat_state_feedback_status = "persisted"
    except HTTPException:
        logger.info("chat_feedback_memory_agent_unavailable")
    except Exception as exc:
        chat_state_feedback_status = "failed"
        logger.warning("chat_feedback_chat_state_persist_failed: %s", exc, exc_info=True)

    learning_actions: List[str] = []
    learning_status = "skipped"
    if (feedback_data.query_text or "").strip() and (feedback_data.response_text or "").strip():
        try:
            from backend.app.core.learning.continuous_learner import get_continuous_learner

            learner = get_continuous_learner()
            learning_result = await learner.process_interaction(
                query=feedback_data.query_text,
                response={
                    "request_id": feedback_data.response_id,
                    "response_text": feedback_data.response_text,
                    "source": feedback_data.source,
                    "confidence": feedback_data.confidence,
                    "mode": feedback_data.mode,
                    "citations": citations,
                },
                feedback_type=feedback_data.feedback_type,
                user_comment=feedback_data.comment,
                confidence_score=feedback_data.confidence,
                session_id=feedback_data.session_id,
                user_id=str(getattr(current_user, "id", "anonymous")),
            )
            learning_actions = list(learning_result.get("actions_taken", []) or [])
            learning_status = "processed"
        except Exception as learner_error:
            learning_status = "failed"
            logger.warning("Falha ao processar feedback no continuous learner: %s", learner_error, exc_info=True)

    return {
        "message": "Feedback submitted successfully.",
        "request_id": feedback_data.response_id,
        "chat_state_feedback_status": chat_state_feedback_status,
        "learning_status": learning_status,
        "learning_actions": learning_actions,
    }


@router.post("/automation/approve", response_class=ORJSONResponse)
async def approve_chat_automation(
    payload: AutomationApproveRequest,
    current_user: Annotated[User, Depends(get_current_active_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    require_chat_capability(current_user, "computer_use")
    if chat_service_v3 is None or session_manager is None:
        await initialize_agents_async()

    proposal = payload.proposal.model_dump() if payload.proposal is not None else None
    automation = await chat_automation_service.approve(
        db,
        current_user=current_user,
        surface="chat",
        proposal=proposal,
        approval_id=payload.approval_id,
        follow_up_action=payload.follow_up_action,
    )
    _persist_automation_state_to_history(automation, current_user)
    return {"automation": automation}


@router.post("/automation/reject", response_class=ORJSONResponse)
async def reject_chat_automation(
    payload: AutomationRejectRequest,
    current_user: Annotated[User, Depends(get_current_active_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
) -> dict:
    require_chat_capability(current_user, "computer_use")
    if chat_service_v3 is None or session_manager is None:
        await initialize_agents_async()

    proposal = payload.proposal.model_dump() if payload.proposal is not None else None
    automation = await chat_automation_service.reject(
        db,
        current_user=current_user,
        surface="chat",
        proposal=proposal,
        approval_id=payload.approval_id,
    )
    _persist_automation_state_to_history(automation, current_user)
    return {"automation": automation}


@router.get("/automation/history", response_class=ORJSONResponse)
async def get_chat_automation_history(
    current_user: Annotated[User, Depends(get_current_active_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
    limit: int = 20,
) -> dict:
    require_chat_capability(current_user, "computer_use")
    items = await chat_automation_service.list_automations(
        db,
        current_user=current_user,
        surface="chat",
        limit=max(1, min(limit, 100)),
    )
    return {"items": items}


@router.get("/automation/artifacts/{approval_id}/{filename}")
async def download_chat_automation_artifact(
    approval_id: str,
    filename: str,
    current_user: Annotated[User, Depends(get_current_active_user)],
    db: Annotated[AsyncSession, Depends(get_db)],
):
    require_chat_capability(current_user, "computer_use")
    automation = await chat_automation_service.get_automation(
        db,
        current_user=current_user,
        surface="chat",
        approval_id=approval_id,
    )
    artifact = automation.get("artifact") if isinstance(automation.get("artifact"), dict) else {}
    if sanitize_text_label(artifact.get("filename"), max_length=180) != sanitize_text_label(filename, max_length=180):
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Artefato não encontrado.")
    artifact_path = chat_automation_service.resolve_artifact_path(approval_id, filename)
    media_type = str(artifact.get("mime_type") or "application/octet-stream")
    return FileResponse(path=artifact_path, media_type=media_type, filename=artifact_path.name)


@router.get("/capabilities", response_class=ORJSONResponse)
async def get_chat_capabilities(
    current_user: Annotated[User, Depends(get_current_active_user)],
    debug: bool = False,
    role: Optional[str] = None,
    username: Optional[str] = None,
    email: Optional[str] = None,
    user_id: Optional[str] = None,
) -> dict:
    target_user: User | SimpleNamespace = current_user
    simulation_requested = any(
        value not in (None, "")
        for value in (role, username, email, user_id)
    )
    if simulation_requested:
        if str(getattr(current_user, "role", "") or "").strip().lower() != "admin":
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail="Somente administradores podem simular capability matrix.",
            )
        target_user = SimpleNamespace(
            id=user_id or getattr(current_user, "id", ""),
            username=username or getattr(current_user, "username", ""),
            email=email or getattr(current_user, "email", ""),
            role=role or getattr(current_user, "role", ""),
        )

    capabilities = get_chat_capabilities_for_user(target_user)
    payload = {
        "capabilities": capabilities,
        "role": str(getattr(target_user, "role", "") or ""),
        "subject": {
            "mode": "simulation" if simulation_requested else "current_user",
            "user_id": str(getattr(target_user, "id", "") or ""),
            "username": str(getattr(target_user, "username", "") or ""),
            "email": str(getattr(target_user, "email", "") or ""),
        },
    }
    if debug:
        payload["diagnostics"] = get_chat_capability_diagnostics_for_user(target_user)
    return payload


@router.post("", response_class=ORJSONResponse)
async def send_chat_message(
    request: ChatRequest,
    current_user: Annotated[User, Depends(get_current_active_user)],
) -> dict:
    from backend.app.core.context import set_current_user_context

    logger.warning("Legacy chat endpoint used.")
    if not _is_chat_allowed_for_user(current_user):
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="ChatBI em canary fechado para este perfil. Solicite liberacao do acesso.",
        )

    if chat_service_v3 is None:
        logger.info("Lazy initializing chat service for POST /chat endpoint")
        await initialize_agents_async()

    if chat_service_v3 is None:
        raise HTTPException(status_code=500, detail="Servico de chat ainda nao inicializado.")

    set_current_user_context(current_user)
    user_capabilities = get_chat_capabilities_for_user(current_user)
    session_id = str(request.session_id or uuid4())
    effective_query = _build_guided_chat_query(
        request.query,
        request.chat_mode,
        request.playbook_context,
        request.guided_action,
    )

    result = await chat_service_v3.process_message(
        query=effective_query,
        session_id=session_id,
        user_id=current_user.id,
        user_role=current_user.role,
        user_capabilities=user_capabilities,
    )
    return {
        "response": str(result),
        "full_agent_response": result,
        "session_id": session_id,
    }


@router.get("/history")
async def get_chat_history(
    current_user: Annotated[User, Depends(get_current_active_user)],
    session_id: Optional[str] = None,
    limit: int = 20,
    offset: int = 0,
) -> dict:
    """
    Recupera histórico persistido da sessão atual e lista sessões do usuário.
    """
    global session_manager
    require_chat_capability(current_user, "memory")

    if session_manager is None:
        await initialize_agents_async()

    if session_manager is None:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Serviço de histórico indisponível no momento.",
        )

    user_id = str(getattr(current_user, "id", ""))
    try:
        sessions = session_manager.list_sessions(user_id=user_id, limit=limit, offset=offset)
        items: List[Dict[str, Any]] = []
        if session_id:
            items = session_manager.get_full_history(session_id=session_id, user_id=user_id)
        return {
            "items": items,
            "sessions": sessions,
            "session_id": session_id,
            "user": current_user.username,
            "capabilities": get_chat_capabilities_for_user(current_user),
        }
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc


@router.delete("/history/{session_id}")
async def delete_chat_history(
    session_id: str,
    current_user: Annotated[User, Depends(get_current_active_user)],
) -> dict:
    """Exclui uma conversa persistida do usuário atual."""
    global session_manager
    require_chat_capability(current_user, "memory")

    if session_manager is None:
        await initialize_agents_async()

    if session_manager is None:
        raise HTTPException(
            status_code=status.HTTP_503_SERVICE_UNAVAILABLE,
            detail="Serviço de histórico indisponível no momento.",
        )

    try:
        session_manager.clear_session(session_id=session_id, user_id=str(getattr(current_user, "id", "")))
        return {"success": True, "session_id": session_id}
    except ValueError as exc:
        raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=str(exc)) from exc
